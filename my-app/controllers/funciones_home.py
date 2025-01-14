
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD
import datetime
import re
import os
from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio
import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file
import socket # biblioteca o modulo para conectar a impresora
from flask import Flask, session, request, redirect, url_for, render_template, flash

# def actualizar_estado_impresion(id_impresora,correlativo_actual, nuevo_correlativo):
#     try:
#         with connectionBD() as conexion_sql:
#             with conexion_sql.cursor() as cursor:
#                 # Actualizar el estado de la impresión a 'finalizado'
#                 querySQL = "UPDATE APP_PALLET.COLA_IMPRESION SET id_estado_impresion = '3' WHERE correlativo_inicial =  ? and correlativo_final= ?"
#                 cursor.execute(querySQL, (correlativo_actual, nuevo_correlativo))
#                 conexion_sql.commit()
#                 print("Estado de impresión actualizado correctamente.")
#                 # Actualizar el estado de la impresión a 'finalizado'
#                 updatesql = "UPDATE APP_PALLET.IMPRESORAS SET connecion_status_id = '1' WHERE id_impresora= ? "
#                 cursor.execute(updatesql, (id_impresora,))
#                 conexion_sql.commit()
#     except Exception as e:
#         print(f"Error al actualizar el estado de la impresión: {str(e)}")



# def valores_para_imprimir_zpl(id_impresora, id_pallet, id_tipo_plantilla,correlativo_actual,nuevo_correlativo):
#     try:
#         with connectionBD() as conexcion_sql:
#             with conexcion_sql.cursor() as cursor:
#                 print("Valor de id_impresion:", correlativo_actual)
#                 print("Valor de id_impresion:", nuevo_correlativo)
#                 print("Valor de id_impresion:", id_pallet)
#                 print("Valor de id_impresion:", id_tipo_plantilla)
#                 print("Valor de id_impresion:", id_impresora)
#                 # traer el printer_ip
#                 querySQL = "SELECT ip_impresora FROM APP_PALLET.IMPRESORAS WHERE id_impresora= ?"
#                 cursor.execute(querySQL, (id_impresora,))
#                 printer = cursor.fetchone()
#                 printer_ip = printer['ip_impresora']
#                 # traer en zpl_template y titulo
#                 querySQL = "SELECT codigo_zpl, titulo_plantilla FROM APP_PALLET.PLANTILLA_ZPL WHERE id_plantilla_pallet = ? AND id_plantilla_tipo = ?"
#                 cursor.execute(querySQL, (id_pallet, id_tipo_plantilla))
#                 plantilla_info = cursor.fetchone()
#                 zpl_template = plantilla_info['codigo_zpl']
#                 titulo = plantilla_info['titulo_plantilla']
#                 print("titulo es:", titulo)
#                 # traer correlativo inicial y correlativo final
#                 carton_inicial = int(correlativo_actual) 
#                 carton_final = int(nuevo_correlativo)
#                 print("Valor de carton_inicial:", carton_inicial)
#                 print("Valor de carton_final:", carton_final)
#                 return printer_ip, 9100, zpl_template, titulo, carton_inicial, carton_final
#     except Exception as e:
#         return f'Se produjo un error en procesar_nueva_impresion: {str(e)}'


import socket
import threading

# # Estructura para almacenar información de hilos
# info_hilos = {}

# def imprimir_zpl(id_impresora, id_pallet, id_tipo_plantilla, correlativo_actual, nuevo_correlativo):
#     printer_ip, printer_port, zpl_template, titulo, carton_inicial, carton_final = valores_para_imprimir_zpl(id_impresora, id_pallet, id_tipo_plantilla, correlativo_actual, nuevo_correlativo)
#     # Crear un socket TCP/IP
#     printer_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

#     try:
#         # Conectar al dispositivo de impresión
#         printer_socket.connect((printer_ip, printer_port))

#         # Iterar sobre el rango de cartones
#         for carton_nbr in range(int(carton_inicial), int(carton_final) + 1):
#             # Reemplazar variables en la plantilla
#             zpl_data = zpl_template.replace('{{titulo}}', titulo).replace('{{carton_nbr}}', str(carton_nbr))

#             # Verificar si se debe detener el hilo
#             if info_hilos.get(id_impresora, {}).get('detener_hilo', False):
#                 print(f'Se detuvo la impresión con id {id_impresora}')
#                 break  # Salir del bucle si se debe detener el hilo

#             # Enviar el comando ZPL a la impresora
#             printer_socket.sendall(zpl_data.encode('utf-8'))
#             print(f"Comando ZPL para cartón {carton_nbr} enviado correctamente a la impresora en la dirección:", printer_ip)
#         # Actualizar el estado de la impresión después de imprimir todos los cartones
#         actualizar_estado_impresion(id_impresora, correlativo_actual, nuevo_correlativo)
#     except socket.error as e:
#         print("Error al intentar conectarse a la impresora:", e)
#     finally:
#         # Cerrar el socket después de enviar todos los comandos
#         printer_socket.close()


# primer modulo 
def actualizar_estado_impresoras(id_impresora):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "SELECT connecion_status_id FROM APP_PALLET.IMPRESORAS WHERE id_impresora = ?"
                cursor.execute(querySQL, (id_impresora,))
                result_impresora = cursor.fetchone()
                if result_impresora is not None:
                    connection_status_id = int(result_impresora[0])
                    impresoratotal = 1 if connection_status_id == 2 else 2
                    return impresoratotal
                else:
                    print("No se encontró ninguna fila para el id_impresora proporcionado.")
                    return None
    except Exception as e:
        print(f"Error en la función calcular_nuevo_correlativo: {e}")
        return None


def calcular_nuevo_correlativo(id_pallet, monto_imprimir):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "SELECT correlativo_actual FROM APP_PALLET.SERIE_PALLET WHERE id_pallet = ?"
                cursor.execute(querySQL, (id_pallet,))
                result = cursor.fetchone()   
                if result is not None:
                    correlativo_actual = result[0]
                    nuevo_correlativo = int(correlativo_actual) + int(monto_imprimir)
                    return int(correlativo_actual), nuevo_correlativo  # Modificado para devolver el correlativo actual también
                else:
                    print("No se encontró ninguna fila para el id_pallet proporcionado.")
                    return None, None
    except Exception as e:
        print(f"Error en la función calcular_nuevo_correlativo: {e}")
        return None, None

def procesar_nueva_impresion(dataForm):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                # Calcula el nuevo correlativo usando la función
                correlativo_actual, nuevo_correlativo = calcular_nuevo_correlativo(dataForm['id_pallet'], dataForm['monto_imprimir'])
                if correlativo_actual is None or nuevo_correlativo is None:
                    print("Error al calcular el nuevo correlativo.")
                    return None, None, None, None

                impresoratotal = actualizar_estado_impresoras(dataForm['id_impresora'])
                
                # Actualiza el correlativo en serie_pallet
                update_sql = "UPDATE APP_PALLET.SERIE_PALLET SET correlativo_actual = ? WHERE id_pallet = ?"
                cursor.execute(update_sql, (str(nuevo_correlativo), dataForm['id_pallet']))

                # Actualiza la impresora
                update_sql = "UPDATE APP_PALLET.IMPRESORAS SET connecion_status_id = '2' WHERE id_impresora = ?"
                cursor.execute(update_sql, dataForm['id_impresora'])
                
                # Inserta en APP_PALLET.COLA_IMPRESION
                insert_sql = """
                    INSERT INTO APP_PALLET.COLA_IMPRESION
                    (id_impresora, id_pallet, id_tipo_plantilla, monto_imprimir, id_estado_impresion, correlativo_inicial, correlativo_final)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                valores = (dataForm['id_impresora'], dataForm['id_pallet'], dataForm['id_tipo_plantilla'], dataForm['monto_imprimir'], str(impresoratotal), str(correlativo_actual), str(nuevo_correlativo))
                cursor.execute(insert_sql, valores)
                conexcion_sql.commit()
                
                resultado_insert = cursor.rowcount
                id_impresion = obtenerid(correlativo_actual, nuevo_correlativo)
                return resultado_insert, correlativo_actual, nuevo_correlativo, id_impresion
    except Exception as e:
        print(f"Se produjo un error en procesar_nueva_impresion: {str(e)}")
        return None, None, None, None


def obtenerid(correlativo_actual, nuevo_correlativo):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "SELECT id_impresion FROM APP_PALLET.COLA_IMPRESION WHERE correlativo_inicial = ? AND correlativo_final = ?"
                cursor.execute(querySQL, (correlativo_actual, nuevo_correlativo))
                resultado = cursor.fetchone()
                if resultado is not None:
                    id_impresion = resultado[0]
                    return id_impresion
                else:
                    print("No se encontró ninguna ID_IMPRESION con los correlativos proporcionados.")
                    return None
    except Exception as e:
        print(f"Error en la función obtenerid: {e}")
        return None
# def procesar_nueva_impresion(dataForm):
#     try:
#         with connectionBD() as conexcion_sql:
#             with conexcion_sql.cursor() as cursor:
#                 # Calcula el nuevo correlativo usando la función
#                 correlativo_actual, nuevo_correlativo  = calcular_nuevo_correlativo(dataForm['id_pallet'],dataForm['monto_imprimir'])
#                 impresoratotal= actualizar_estado_impresoras(dataForm['id_impresora'])
#                 # Actualiza el correlativo en serie_pallet
#                 update_sql = "UPDATE APP_PALLET.SERIE_PALLET SET correlativo_actual = ? WHERE id_pallet = ?"
#                 cursor.execute(update_sql, (str(nuevo_correlativo), dataForm['id_pallet']))
    
#                 # Actualiza la impresora
#                 update_sql = "UPDATE APP_PALLET.IMPRESORAS  SET connecion_status_id = '2' WHERE id_impresora = ?"
#                 cursor.execute(update_sql,dataForm['id_impresora'])
#                 # Primera consulta: INSERT INTO APP_PALLET.COLA_IMPRESION
#                 insert_sql = "INSERT INTO APP_PALLET.COLA_IMPRESION (id_impresora, id_pallet, id_tipo_plantilla, monto_imprimir,id_estado_impresion,correlativo_inicial,correlativo_final) VALUES (?, ?, ?, ?,?,?, ?)"
#                 valores = (dataForm['id_impresora'], dataForm['id_pallet'], dataForm['id_tipo_plantilla'], dataForm['monto_imprimir'],str(impresoratotal),str(correlativo_actual),str(nuevo_correlativo))
#                 cursor.execute(insert_sql, valores)
#                 conexcion_sql.commit()
#                 resultado_insert = cursor.rowcount
#                 id_impresion= obtenerid(correlativo_actual,nuevo_correlativo)
#                 return resultado_insert,correlativo_actual,nuevo_correlativo,id_impresion
#     except Exception as e:
#         return f'Se produjo un error en procesar_nueva_impresion: {str(e)}'


# def obtenerid(correlativo_actual,nuevo_correlativo):
#     try:
#         with connectionBD() as conexcion_sql:
#             with conexcion_sql.cursor() as cursor:
#                 querySQL = "SELECT id_impresion  FROM APP_PALLET.COLA_IMPRESION  WHERE correlativo_inicial =  ? and correlativo_final= ?"
#                 cursor.execute(querySQL, (correlativo_actual, nuevo_correlativo))
#                 resultado = cursor.fetchone()
#                 if resultado is not None:
#                     id_impresion =resultado['id_impresion']
#                     return id_impresion
#                 else:
#                     print("No se encontró ninguna ID_IMPRESION  proporcionado.")
#                     return None
#     except Exception as e:
#         print(f"Error en la función calcular_nuevo_correlativo: {e}")
#         return None




# segundo modulo - lista cola de imprecion 
def sql_lista_cola_impresiones():
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_impresion,
                        e.monto_imprimir,
                        s.descripcion,
                        i.nombre_impresora,
						w.nombre_status as nombre_status2,
                        t.nombre_status,
                        p.formato
                    FROM APP_PALLET.COLA_IMPRESION AS e
                    INNER JOIN APP_PALLET.IMPRESORAS i on e.id_impresora = i.id_impresora
					INNER JOIN APP_PALLET.SERIE_PALLET s on   e.id_pallet=s.id_pallet
					INNER JOIN APP_PALLET.STATUS_IMPRESORA t on i.connecion_status_id =t.connecion_status_id
					INNER JOIN APP_PALLET.STATUS_IMPRESION w on e.id_estado_impresion=w.id_estado_impresion
     				INNER JOIN APP_PALLET.TIPO_PLANTILLA p on  e.id_tipo_plantilla = p.id_tipo_plantilla
                    ORDER BY e.id_impresion DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función lista_cola_impresiones: {e}")
        return None


# detalle de impresion  
def sql_detalle_impresion(idImpresion):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_impresion,
                        e.monto_imprimir,
                        s.descripcion,
                        i.nombre_impresora,
						w.nombre_status as nombre_status2,
                        t.nombre_status,
						e.correlativo_inicial,
                        p.formato,
						e.correlativo_final
                    FROM APP_PALLET.COLA_IMPRESION AS e
                    INNER JOIN APP_PALLET.IMPRESORAS i on e.id_impresora = i.id_impresora
					INNER JOIN APP_PALLET.SERIE_PALLET s on   e.id_pallet=s.id_pallet
					INNER JOIN APP_PALLET.STATUS_IMPRESORA t on i.connecion_status_id =t.connecion_status_id
					INNER JOIN APP_PALLET.STATUS_IMPRESION w on e.id_estado_impresion=w.id_estado_impresion
                    INNER JOIN APP_PALLET.TIPO_PLANTILLA p on  e.id_tipo_plantilla = p.id_tipo_plantilla
                    WHERE e.id_impresion =?
                    ORDER BY e.id_impresion DESC
                    """)
                cursor.execute(querySQL, (idImpresion,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalle_impresion: {e}")
        return None

# buscador por impresora 
def buscarimpresora(search):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_impresion,
                            e.monto_imprimir,
                            s.descripcion,
                            i.nombre_impresora,
                            w.nombre_status as nombre_status2,
                            t.nombre_status,
                            p.formato
                        FROM APP_PALLET.COLA_IMPRESION AS e
                        INNER JOIN APP_PALLET.IMPRESORAS i on e.id_impresora = i.id_impresora
                        INNER JOIN APP_PALLET.SERIE_PALLET s on   e.id_pallet=s.id_pallet
                        INNER JOIN APP_PALLET.STATUS_IMPRESORA t on i.connecion_status_id =t.connecion_status_id
                        INNER JOIN APP_PALLET.STATUS_IMPRESION w on e.id_estado_impresion=w.id_estado_impresion
                        INNER JOIN APP_PALLET.TIPO_PLANTILLA p on  e.id_tipo_plantilla = p.id_tipo_plantilla
                        WHERE i.nombre_impresora LIKE  ?
                        ORDER BY e.id_impresion DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarimpresora: {e}")
        return []

# tercer modulo Funcion Empleados Informe (Reporte)
def impresiones_reporte():
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_impresion,
                        e.monto_imprimir,
                        s.descripcion,
                        i.nombre_impresora,
                        r.nombre_status,
                        e.fecha_registro
                    FROM APP_PALLET.COLA_IMPRESION AS e
                    INNER JOIN APP_PALLET.IMPRESORAS i on e.id_impresora = i.id_impresora
                    INNER JOIN APP_PALLET.STATUS_IMPRESORA r on i.connecion_status_id = r.connecion_status_id
                    INNER JOIN APP_PALLET.SERIE_PALLET s on e.id_pallet = s.id_pallet
                    ORDER BY e.id_impresion DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función impresiones_reporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = impresiones_reporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Id_impresion", "Monto Impreso", "Nombre de correlativo",
                     "Impresora", "Estado de Impresion", "Fecha de Impresion")

    hoja.append(cabeceraExcel)

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        id_impresion = registro['id_impresion']
        monto_imprimir = registro['monto_imprimir']
        descripcion = registro['descripcion']
        nombre_impresora = registro['nombre_impresora']
        nombre_status = registro['nombre_status']
        fecha_registro = registro['fecha_registro']
        # Agregar los valores a la hoja
        hoja.append((id_impresion, monto_imprimir, descripcion, nombre_impresora, nombre_status, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_Impresiones_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


# modulo 4 y5  usuarios 

# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []

# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "DELETE FROM users WHERE id=?"
                cursor.execute(querySQL, (id,))
                conexcion_sql.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
# modulo 3 - lista_correlativos
def lista_correlativos():
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "SELECT id_pallet,serie,descripcion,correlativo_actual FROM APP_PALLET.SERIE_PALLET "
                cursor.execute(querySQL,)
                usuariosBD2 = cursor.fetchall()
        return usuariosBD2
    except Exception as e:
        print(f"Error en lista_correlativos : {e}")
        return []


# Eliminar uEmpleado
def eliminarImpresion(id_impresion):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = "DELETE FROM APP_PALLET.COLA_IMPRESION  WHERE id_impresion=?"
                cursor.execute(querySQL, (id_impresion,))
                conexcion_sql.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# reemprimir 
def reemprimir(id_impresion):
    try:
        with connectionBD() as conexcion_sql:
            with conexcion_sql.cursor() as cursor:
                querySQL = ("""
                select  e.id_impresora,i.ip_impresora,r.titulo_plantilla,r.codigo_zpl,e.correlativo_inicial,e.correlativo_final
                from APP_PALLET.COLA_IMPRESION e
                inner join APP_PALLET.IMPRESORAS i on i.id_impresora=e.id_impresora
                inner join APP_PALLET.PLANTILLA_ZPL r on r.id_plantilla_pallet=e.id_pallet and r.id_plantilla_tipo=e.id_tipo_plantilla
                WHERE id_impresion=?""")
                cursor.execute(querySQL, (id_impresion,))
                plantilla_reemprimir = cursor.fetchone()
                printer_ip = plantilla_reemprimir['ip_impresora']
                titulo = plantilla_reemprimir['titulo_plantilla']
                zpl_template = plantilla_reemprimir['codigo_zpl']
                carton_inicial = plantilla_reemprimir['correlativo_inicial']
                carton_final = plantilla_reemprimir['correlativo_final']
                id_impresora = plantilla_reemprimir['id_impresora']
        return printer_ip, 9100, zpl_template, titulo, carton_inicial, carton_final,id_impresora
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []

def reimpimir_zpl(id_impresion):
    try:
        # Obtener información necesaria para la reimpresión
        printer_ip, printer_port, zpl_template, titulo, carton_inicial, carton_final, id_impresora = reemprimir(id_impresion)

        # Crear un socket TCP/IP
        printer_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

        try:
            # Conectar al dispositivo de impresión
            printer_socket.connect((printer_ip, printer_port))

            # Variable de control para detener el hilo
            detener_hilo = False

            # Iterar sobre el rango de cartones
            for carton_nbr in range(int(carton_inicial), int(carton_final) + 1):
                # Verificar si se ha solicitado detener el hilo
                if detener_hilo:
                    break

                # Reemplazar variables en la plantilla
                zpl_data = zpl_template.replace('{{titulo}}', titulo).replace('{{carton_nbr}}', str(carton_nbr))

                # Enviar el comando ZPL a la impresora
                printer_socket.sendall(zpl_data.encode('utf-8'))
                print(f"Comando ZPL para cartón {carton_nbr} enviado correctamente a la impresora en la dirección:", printer_ip)

                # Simular un tiempo de espera entre impresiones
                threading.Event().wait(5)

            # Actualizar el estado de la impresión después de imprimir todos los cartones
            actualizar_estado_reeimpresion(id_impresion, id_impresora)
        except socket.error as e:
            print("Error al intentar conectarse a la impresora:", e)
        finally:
            # Cerrar el socket después de enviar todos los comandos
            printer_socket.close()
    except Exception as e:
        print(f"Error en la reimpresión en segundo plano: {e}")
        flash('Se produjo un error durante la reimpresión', 'error')

    # Eliminar el hilo de la estructura
    del info_hilos[id_impresion]


def actualizar_estado_reeimpresion(id_impresion,id_impresora):
    try:
        with connectionBD() as conexion_sql:
            with conexion_sql.cursor() as cursor:
                # Actualizar el estado de la impresión a 'finalizado'
                querySQL = "UPDATE APP_PALLET.COLA_IMPRESION SET id_estado_impresion = '4' WHERE id_impresion= ?"
                cursor.execute(querySQL, (id_impresion,))
                conexion_sql.commit()
                print("Estado de impresión actualizado correctamente.")
                # Actualizar el estado de la impresión a 'finalizado'
                updatesql = "UPDATE APP_PALLET.IMPRESORAS SET connecion_status_id = '1' WHERE id_impresora= ? "
                cursor.execute(updatesql, (id_impresora,))
                conexion_sql.commit()
    except Exception as e:
        print(f"Error al actualizar el estado de la impresión: {str(e)}")
        
    
    
        
# Función para detener la impresión de un hilo específico
def detener_impresion(id_impresion):
    if id_impresion in info_hilos:
        # Establecer la variable de control para detener el hilo
        info_hilos[id_impresion]['detener_hilo'] = True
        flash(f'Se detuvo la impresión con id {id_impresion}', 'info')
    else:
        flash(f'No se encontró la impresión con id {id_impresion}', 'error')


# Estructura para almacenar información de hilos
info_hilos = {}


